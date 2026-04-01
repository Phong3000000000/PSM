from odoo import models, fields, api, _
from odoo.exceptions import ValidationError, AccessError, UserError
import os
import logging
import json

_logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:
    _logger.warning("openpyxl not found, generic excel import will fail")
    openpyxl = None


class PifObject(models.Model):
    _name = 'pif.object'
    _description = 'PIF Execution Object'
    _inherit = ['mail.thread', 'mail.activity.mixin']

    def _get_default_tracking_steps(self):
        return [
            ('1_rsg', 'RSG creates PIF', 'RSG'),
            ('2_it_config', 'IT config (check)', 'IT'),
            ('3_master', 'RSG setup', 'RSG'),
            ('4_lab', 'RSG lab test', 'RSG'),
            ('5_pilot', 'Pilot/Mass', 'IT'),
            ('6_val', 'PIF Valuation', 'RSG'),
        ]

    @api.model
    def default_get(self, fields_list):
        defaults = super(PifObject, self).default_get(fields_list)
        if 'process_tracking_ids' in fields_list:
            steps = self._get_default_tracking_steps()
            lines = []
            for code, name, dept in steps:
                lines.append((0, 0, {
                    'step_code': code,
                    'process_name': name,
                    'department_ref': dept,
                    'status': 'pending',
                }))
            defaults['process_tracking_ids'] = lines
        return defaults

    name = fields.Char(string='PIF ID', required=True, copy=False, readonly=True, index=True, default=lambda self: _('PIF Creation'))
    
    
    approval_request_id = fields.Many2one('approval.request', string='Origin Request', required=False, readonly=True)
    
    # Request Info (Related)
    # Request Info
    request_owner_id = fields.Many2one('res.users', string='Request Owner', default=lambda self: self.env.user, readonly=True)
    request_owner_department_id = fields.Many2one('hr.department', string='Department', compute='_compute_owner_info', store=True, readonly=True)
    request_owner_job_id = fields.Many2one('hr.job', string='Job Position', compute='_compute_owner_info', store=True, readonly=True)

    is_request_owner = fields.Boolean(compute='_compute_is_request_owner')
    user_can_process_step = fields.Boolean(compute='_compute_user_can_process_step')
    is_it_dept = fields.Boolean(compute='_compute_is_it_dept')
    is_rsg_dept = fields.Boolean(compute='_compute_is_rsg_dept')

    # IT Config Data

    # IT Config Data
    it_data_line_ids = fields.One2many(
        'pif.data.line', 'pif_object_id', string='IT Config Data',
        domain=[('category', 'in', ['SSBI', 'RFM', 'POS'])]
    )

    def unlink(self):
        if not (self.env.is_superuser() or self.env.user.has_group('base.group_system')):
             raise UserError(_("You cannot delete PIF requests. Please functionalities Refuse/Cancel instead."))
        return super(PifObject, self).unlink()
    
    @api.depends('state')
    def _compute_is_it_dept(self):
        # Check if current user is in IT Dept
        it_dept = False
        employee = self.env['hr.employee'].sudo().search([('user_id', '=', self.env.user.id)], limit=1)
        if employee and employee.department_id and 'IT' in employee.department_id.name: # Assuming 'IT' keyword
             it_dept = True
        
        for rec in self:
            rec.is_it_dept = it_dept

    @api.depends('state')
    def _compute_is_rsg_dept(self):
        rsg_dept = False
        employee = self.env['hr.employee'].sudo().search([('user_id', '=', self.env.user.id)], limit=1)
        if employee and employee.department_id and 'RSG' in employee.department_id.name:
             rsg_dept = True
        for rec in self:
            rec.is_rsg_dept = rsg_dept

    @api.depends('state')
    def _compute_user_can_process_step(self):
        state_map = {
            'rsg_create': 'RSG',
            'it_config': 'IT',
            'master_data': 'RSG',
            'lab_test': 'RSG',
            'pilot': 'IT',
            'valuation': 'RSG',
        }
        
        employee = self.env['hr.employee'].sudo().search([('user_id', '=', self.env.user.id)], limit=1)
        user_dept = employee.department_id.name if employee and employee.department_id else ''
        
        for rec in self:
            required = state_map.get(rec.state)
            
            if not required:
                 rec.user_can_process_step = False
                 continue
                 
            if required == 'Head':
                 creator = rec.request_owner_id
                 is_manager = False
                 
                 creator_emp = self.env['hr.employee'].sudo().search([('user_id', '=', creator.id)], limit=1)
                 if creator_emp:
                     # 1. Direct Parent
                     if creator_emp.parent_id and creator_emp.parent_id.user_id == self.env.user:
                         is_manager = True
                     # 2. Dept Manager
                     elif creator_emp.department_id.manager_id and creator_emp.department_id.manager_id.user_id == self.env.user:
                         is_manager = True
                     # 3. Admin
                     if self.env.is_superuser() or self.env.user.has_group('base.group_system'):
                         is_manager = True
                         
                 rec.user_can_process_step = is_manager
                 
            else:
                 rec.user_can_process_step = required in user_dept
                 
                 if self.env.is_superuser() or self.env.user.has_group('base.group_system'):
                      rec.user_can_process_step = True

    @api.depends('request_owner_id')
    def _compute_is_request_owner(self):
        for rec in self:
            rec.is_request_owner = (rec.request_owner_id == self.env.user)

    @api.depends('request_owner_id')
    def _compute_owner_info(self):
        for rec in self:
            if rec.request_owner_id:
                employee = self.env['hr.employee'].sudo().search([('user_id', '=', rec.request_owner_id.id)], limit=1)
                rec.request_owner_department_id = employee.department_id if employee else False
                rec.request_owner_job_id = employee.job_id if employee else False
            else:
                rec.request_owner_department_id = False
                rec.request_owner_job_id = False
    
    state = fields.Selection([
        ('rsg_create', 'RSG Created'),
        ('it_config', 'IT Configured'),
        ('master_data', 'Master Data Set'),
        ('lab_test', 'Lab Tested'),
        ('pilot', 'Pilot Done'),
        ('valuation', 'Valuation Done'),
        ('approved', 'Approved'),  # Final state after valuation
    ], string='Status', required=True, readonly=True, copy=False, tracking=True, default='rsg_create')

    # Core Info
    pif_request_type = fields.Selection([
        ('menu', 'Menu'),
        ('marketing', 'Marketing'),
        ('si', 'S&I'),
        ('digital', 'Digital'),
        ('supply_chain', 'Supply Chain'),
    ], string='Request Type', default='menu')
    
    pif_product_id = fields.Many2one('product.product', string='Product', required=False) # Optional now
    formula_description = fields.Text(string='Formula Description')
    
    # NEW: General Description
    description_html = fields.Html(string='General Description')

    supplier_info = fields.Text(string='Supplier Information')
    
    # UI Display Fields
    pif_product_name = fields.Char(related='pif_product_id.name', string='Product Name')
    pif_product_code = fields.Char(related='pif_product_id.default_code', string='Code/WRIN')
    pif_product_vendor = fields.Char(compute='_compute_pif_product_vendor', string='Vendor')
    pif_product_status = fields.Selection(related='pif_product_id.pif_status', string='Product Status', readonly=True)
    pif_bom_status = fields.Selection(related='pif_bom_id.pif_status', string='BOM Status', readonly=True)
    
    # BOM & Raw Materials
    # BOM & Raw Materials
    # User Request: Select BOM directly.
    pif_bom_id = fields.Many2one('mrp.bom', string='Select BOM', required=True, domain="[('pif_status', '=', 'approved')]") 
    raw_material_ids = fields.One2many('pif.object.raw.line', 'pif_object_id', string='Raw Materials')
    
    # 3 HTML Inputs Requirement (Replaces Files)
    # 3 HTML Inputs Requirement (Replaces Files)
    file_ssbi_html = fields.Html(string='SSBI Description')
    file_rfm_html = fields.Html(string='RFM Description')
    file_pos_html = fields.Html(string='POS Description')

    @api.onchange('pif_bom_id')
    def _onchange_pif_bom_id(self):
        if not self.pif_bom_id:
             self.raw_material_ids = [(5, 0, 0)]
             return
             
        # Auto-populate Raw Materials from BOM
        bom = self.pif_bom_id
        if bom:
             # Auto-set Product if linked?
             if bom.product_tmpl_id:
                 # Find a variant or just use template name for display
                 pass
                 
             lines = []
             for bom_line in bom.bom_line_ids:
                 lines.append((0, 0, {
                     'gri_code': bom_line.product_id.default_code or '',
                     'wrin_code': bom_line.product_id.product_tmpl_id.wrin_code or '', 
                     'product_id': bom_line.product_id.id,
                     'quantity': bom_line.product_qty,
                     'uom_id': bom_line.product_uom_id.id,
                 }))
             self.raw_material_ids = [(5, 0, 0)] + lines

    @api.depends('pif_product_id')
    def _compute_pif_product_vendor(self):
        for rec in self:
            partners = rec.pif_product_id.seller_ids.mapped('partner_id.name')
            rec.pif_product_vendor = ", ".join(partners) if partners else ""
    
    # Lab Test
    lab_test_result = fields.Selection([
        ('pass', 'Pass'),
        ('fail', 'Fail')
    ], string='Lab Test Result', tracking=True)
    lab_test_note = fields.Text(string='Lab Test Notes')
    lab_test_proof = fields.Binary(string='Lab Test Proof')
    lab_test_proof_filename = fields.Char(string='Proof Filename')
    
    # Pilot Phase
    pilot_store_ids = fields.Many2many('res.partner', string='Target Stores / Pilot Locations', domain="[('is_company', '=', True)]")
    pilot_note = fields.Text(string='Pilot Notes / Result')
    
    # Relations
    data_line_ids = fields.One2many('pif.data.line', 'pif_object_id', string='Data Lines')
    process_tracking_ids = fields.One2many('pif.process.tracking', 'pif_object_id', string='Process Tracking')
    lab_history_ids = fields.One2many('pif.lab.history', 'pif_id', string='Lab History')

    @api.model_create_multi
    def create(self, vals_list):
        # Restriction: Removed (RSG creates)
        for vals in vals_list:
            if vals.get('name', _('PIF Creation')) == _('PIF Creation'):
                 vals['name'] = self.env['ir.sequence'].next_by_code('pif.object') or _('PIF Creation')
        
        res = super(PifObject, self).create(vals_list)
        for record in res:
            if not record.process_tracking_ids:
                record._create_default_tracking()
        return res

    def _create_default_tracking(self):
        steps = self._get_default_tracking_steps()
        lines = []
        for code, name, dept in steps:
            lines.append((0, 0, {
                'step_code': code,
                'process_name': name,
                'department_ref': dept,
                'status': 'pending',
            }))
        self.write({'process_tracking_ids': lines})
        
        # Mark first step as done immediately upon creation? 
        # User said "Menu department staff creates a PIF" is the first step.
        # If created, this is effectively done or pending submission. Let's mark it done when they click Submit.

    def _check_dept_permission(self, dept_name_substr):
        """Helper to enforce Department permissions."""
        # Admin bypass for demo/testing
        if self.env.is_superuser() or self.env.user.has_group('base.group_system'):
            return True
        
        employee = self.env['hr.employee'].sudo().search([('user_id', '=', self.env.user.id)], limit=1)
        if not employee or not employee.department_id or dept_name_substr not in employee.department_id.name:
             raise AccessError(_("You do not have permission. This step requires partcipation from the '%s' Department.") % dept_name_substr)
        return True

    def _complete_step(self, step_prefix):
        """Helper to mark a step as done"""
        line = self.process_tracking_ids.filtered(lambda l: l.step_code.startswith(step_prefix))
        
        # Get User's Department
        employee = self.env['hr.employee'].sudo().search([('user_id', '=', self.env.user.id)], limit=1)
        dept_name = employee.department_id.name if employee and employee.department_id else ''

        if line:
            line.write({
                'status': 'done',
                'user_id': self.env.user.id,
                'department_ref': dept_name or line.department_ref, 
                'execution_date': fields.Datetime.now(),
            })

    # Workflow Actions
    # action_submit_request REMOVED (Draft phase removed)

    def action_rsg_create(self):
        """RSG performs Step 1 -> Moves to IT."""
        self._check_dept_permission('RSG')
        
        # Validation: 3 HTML Fields Required (RSG Entry)
        if not (self.file_ssbi_html and self.file_rfm_html and self.file_pos_html):
             raise ValidationError(_("Please enter descriptions for all 3 required sections (SSBI, RFM, POS) before completing this step."))

        self.action_import_data_from_excel()
        
        self.write({'state': 'it_config'})
        self._complete_step('1_') 
        
        # Notify IT
        it_dept = self.env['hr.department'].search([('name', '=', 'IT')], limit=1)
        if not it_dept: it_dept = self.env['hr.department'].search([('name', 'ilike', 'IT %')], limit=1)
        if it_dept and it_dept.manager_id and it_dept.manager_id.user_id:
             self.activity_schedule(
                'mail.mail_activity_data_todo',
                user_id=it_dept.manager_id.user_id.id,
                summary=_('PIF Configuration Required'),
                note=_('RSG has created the PIF. Please proceed with System Configuration.'),
            )

    def action_it_config(self):
        """IT performs Step 2 -> Moves to RSG (Master)."""
        self._check_dept_permission('IT')
        
        self.write({'state': 'master_data'})
        self._complete_step('2_')
        
        # Notify RSG
        rsg_dept = self.env['hr.department'].search([('name', 'ilike', 'RSG')], limit=1)
        if rsg_dept and rsg_dept.manager_id and rsg_dept.manager_id.user_id:
             self.activity_schedule(
                'mail.mail_activity_data_todo',
                user_id=rsg_dept.manager_id.user_id.id,
                summary=_('System Setup Required'),
                note=_('IT Config Done. Please proceed with Master Data Setup.'),
            )

    def action_master_setup(self):
        """RSG performs Step 3 -> Moves to Lab Test (RSG)."""
        self._check_dept_permission('RSG')
        
        self.write({'state': 'lab_test'}) 
        self._complete_step('3_') 
        
        # Notify RSG (Lab Test)
        rsg_dept = self.env['hr.department'].search([('name', 'ilike', 'RSG')], limit=1)
        user_to_notify = rsg_dept.manager_id.user_id.id if rsg_dept and rsg_dept.manager_id else self.env.user.id
         
        self.activity_schedule(
            'mail.mail_activity_data_todo',
            user_id=user_to_notify,
            summary=_('Lab Test Required'),
            note=_('Master Data Setup Done. Please proceed with Lab Testing.'),
        )
        
    def action_finish_lab_test(self):
        """Unified Action for Lab Test Approve button."""
        self._check_dept_permission('RSG')
        
        if not self.lab_test_result:
            raise ValidationError(_("Please select a Lab Test Result (Pass/Fail) before approving."))

        # LOG HISTORY (Always log regardless of Pass/Fail)
        self.env['pif.lab.history'].create({
            'pif_id': self.id,
            'result': self.lab_test_result,
            'note': self.lab_test_note,
            'proof_file': self.lab_test_proof,
            'proof_filename': self.lab_test_proof_filename,
        })

        if self.lab_test_result == 'pass':
             # Logic for PASS
             self.write({'state': 'pilot'})
             self._complete_step('4_') # 4. Lab Test Done (RSG lab test)
             
             it_dept = self.env['hr.department'].search([('name', '=', 'IT')], limit=1)
             if not it_dept: it_dept = self.env['hr.department'].search([('name', 'ilike', 'IT %')], limit=1)
             user_to_notify = it_dept.manager_id.user_id.id if it_dept and it_dept.manager_id else self.env.user.id
             
             self.activity_schedule(
                'mail.mail_activity_data_todo',
                user_id=user_to_notify,
                summary=_('Ready for Pilot'),
                note=_('Lab Test Passed. Please proceed with Pilot Deployment.'),
            )
            
        elif self.lab_test_result == 'fail':
             # Logic for FAIL
             if not self.lab_test_note:
                 raise ValidationError(_("Please provide a Message/Note explaining the failure."))
                 
             # Return to State 'it_config' for IT to re-check
             self.write({'state': 'it_config'})
              
             # Reset Steps: 2 (IT), 3 (Master), 4 (Lab), 5 (Pilot - pending anyway)
             steps_to_reset = ['2_it_config', '3_master', '4_lab', '5_pilot']
             lines_to_reset = self.process_tracking_ids.filtered(lambda l: l.step_code in steps_to_reset)
             lines_to_reset.write({'status': 'pending', 'user_id': False, 'execution_date': False})
             
             # Notify IT
             it_dept = self.env['hr.department'].search([('name', '=', 'IT')], limit=1)
             if not it_dept: it_dept = self.env['hr.department'].search([('name', 'ilike', 'IT %')], limit=1)
             user_to_notify = it_dept.manager_id.user_id.id if it_dept and it_dept.manager_id else self.env.user.id
    
             self.activity_schedule(
                'mail.mail_activity_data_todo',
                user_id=user_to_notify,
                summary=_('Lab Test Failed - Revision Required'),
                note=_('Lab Test Failed. Please check notes and update IT Config.'),
            )
            
        # Clear Input Fields after saving to history
        self.write({
            'lab_test_result': False,
            'lab_test_note': False,
            'lab_test_proof': False,
            'lab_test_proof_filename': False,
        })
        
    def action_pilot_deploy(self):
        """IT performs Step 5 -> Moves to Valuation."""
        self._check_dept_permission('IT')
        self.write({'state': 'valuation'}) 
        self._complete_step('5_') 
        
        # Notify RSG
        rsg_dept = self.env['hr.department'].search([('name', 'ilike', 'RSG')], limit=1)
        user_to_notify = rsg_dept.manager_id.user_id.id if rsg_dept and rsg_dept.manager_id else self.env.user.id
         
        self.activity_schedule(
            'mail.mail_activity_data_todo',
            user_id=user_to_notify,
            summary=_('Valuation Report Required'),
            note=_('Pilot Deployment Done. Please proceed with Valuation Report.'),
        )

    def action_valuation_report(self):
        """Final approval - generates Finished Good WRIN for the product."""
        self._check_dept_permission('RSG')
        self.write({'state': 'approved'})  # Final State
        self._complete_step('8_')
        
        # Generate Finished Good WRIN
        self._generate_finished_wrin()

    def _generate_finished_wrin(self):
        """
        Generate Finished Good WRIN when PIF is approved.
        
        Input: pif.object (linked to BOM and Product)
        Output: WRIN code on product.template (wrin_code and default_code)
        
        WRIN format: FG-YYYY-XXXX (e.g., FG-2026-0001)
        """
        self.ensure_one()
        
        # Get product template from BOM or product
        product_tmpl = None
        if self.pif_bom_id:
            product_tmpl = self.pif_bom_id.product_tmpl_id
        elif self.pif_product_id:
            product_tmpl = self.pif_product_id.product_tmpl_id
        
        if not product_tmpl:
            return
        
        # Generate WRIN if not exists
        if not product_tmpl.wrin_code or not product_tmpl.default_code:
            # Use sequence or generate new code
            new_wrin = self.env['ir.sequence'].next_by_code('product.wrin.finished') or \
                       f"FG-{fields.Date.today().year}-{product_tmpl.id:04d}"
            
            # Update product template
            product_tmpl.sudo().write({
                'wrin_code': new_wrin,
                'default_code': new_wrin,
                'pif_status': 'approved',
            })
            
            # Also update BOM pif_status
            if self.pif_bom_id:
                self.pif_bom_id.sudo().write({'pif_status': 'approved'})
            
            # Log the WRIN creation
            self.message_post(
                body=f"WRIN Thành phẩm đã được tạo: {new_wrin}",
                message_type='notification',
            )

    def action_import_data_from_excel(self):
        """Reads specific Excel file from file/ folder and populates data lines."""
        # Check standard Docker mount path first
        file_dir = '/mnt/file'
        if not os.path.exists(file_dir):
            # Fallback to relative path (for local dev without docker)
            base_path = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
            file_dir = os.path.join(base_path, 'file')
        target_file = None
        
        # Hardcoded check for the file mentioned or generic search
        if os.path.exists(file_dir):
            for fname in os.listdir(file_dir):
                if fname.endswith('.xlsx') and 'PIF' in fname:
                    target_file = os.path.join(file_dir, fname)
                    break
                    
        if not target_file or not os.path.exists(target_file):
             # Notify user via warning in logs or create a dummy error line
             self.env['bus.bus']._sendone(self.env.user.partner_id, 'simple_notification', {
                'title': _('File Not Found'),
                'message': _('Could not find the PIF Excel file in file/ folder.'),
                'type': 'danger',
            })
             return

        if not openpyxl:
            return

        wb = openpyxl.load_workbook(target_file, data_only=True, read_only=True)
        
        data_vals = []
        # Mapping Sheet Name -> Category Type
        sheet_map = {
            'SSBI': 'ssbi',
            'RFM': 'rfm',
            'RFM 2508': 'rfm',
            'RFM 0909': 'rfm',
            'POS layout': 'pos',
            'POS layout 0909': 'pos',
            'POS layout 08112024': 'pos'
        }

        # Clear existing? Maybe not.
        
        for sheet_name in wb.sheetnames:
            category = sheet_map.get(sheet_name)
            if not category:
                # Try partial match
                if 'SSBI' in sheet_name: category = 'ssbi'
                elif 'RFM' in sheet_name: category = 'rfm'
                elif 'POS' in sheet_name: category = 'pos'
                elif 'MDS' in sheet_name: category = 'mds'
                elif 'UDP' in sheet_name: category = 'udp'
            
            if category:
                ws = wb[sheet_name]
                
                rows = list(ws.iter_rows(values_only=True))
                if not rows: continue
                
                headers = rows[0]
                for i, row in enumerate(rows[1:], start=2):
                    for j, cell_value in enumerate(row):
                         if cell_value:
                             key = f"{headers[j] if j < len(headers) and headers[j] else 'Col'+str(j)} (R{i})"
                             value = str(cell_value)
                             data_vals.append({
                                 'category': category,
                                 'key': key,
                                 'value': value,
                             })

        if data_vals:
            self.data_line_ids = [(0, 0, v) for v in data_vals]



class PifDataLine(models.Model):
    _name = 'pif.data.line'
    _description = 'PIF Command Data Line'

    pif_object_id = fields.Many2one('pif.object', string='PIF Object', ondelete='cascade')
    category = fields.Selection([
        ('ssbi', 'SSBI'),
        ('rfm', 'RFM'),
        ('pos', 'POS'),
        ('mds', 'MDS'),
        ('udp', 'UDP'),
        ('access', 'Access & Audit'),
        ('other', 'Other'),
    ], string='System/Category', required=True, default='other')
    key = fields.Char(string='Command/Parameter', required=True)
    value = fields.Text(string='Value/Config')


class PifObjectRawLine(models.Model):
    _name = 'pif.object.raw.line'
    _description = 'PIF Object Raw Material'

    pif_object_id = fields.Many2one('pif.object', string='PIF Object', ondelete='cascade')
    
    gri_code = fields.Char(string='GRI')
    wrin_code = fields.Char(string='WRIN')
    product_id = fields.Many2one('product.product', string='Raw Item Name')
    quantity = fields.Float(string='Quantity')
    uom_id = fields.Many2one('uom.uom', string='UoM')


class PifProcessTracking(models.Model):
    _name = 'pif.process.tracking'
    _description = 'PIF Process Tracking Step'
    _order = 'id'

    pif_object_id = fields.Many2one('pif.object', string='PIF Object', ondelete='cascade')
    step_code = fields.Char(string='Step Code', required=True)
    
    process_name = fields.Char(string='Process Name')
    department_ref = fields.Char(string='Department') 
    
    status = fields.Selection([
        ('pending', 'Pending'),
        ('processing', 'Processing'),
        ('done', 'Done'),
        ('fail', 'Failed'),
    ], string='Status', default='pending')
    
    user_id = fields.Many2one('res.users', string='Executed By')
    execution_date = fields.Datetime(string='Execution Date')

class PifLabHistory(models.Model):
    _name = 'pif.lab.history'
    _description = 'PIF Lab Test History'
    _order = 'test_date desc'

    pif_id = fields.Many2one('pif.object', string='PIF', ondelete='cascade')
    test_date = fields.Datetime(string='Test Date', default=fields.Datetime.now, readonly=True)
    tester_id = fields.Many2one('res.users', string='Tester', default=lambda self: self.env.user, readonly=True)
    result = fields.Selection([
        ('pass', 'Pass'),
        ('fail', 'Fail')
    ], string='Result', readonly=True)
    note = fields.Text(string='Note', readonly=True)
    proof_file = fields.Binary(string='Attachment', readonly=True)
    proof_filename = fields.Char(string='Filename', readonly=True)
