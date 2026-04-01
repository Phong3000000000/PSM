# -*- coding: utf-8 -*-
"""
SOC Excel Import Wizard – M02_P0209_02
Supports uploading a single .xlsx file (eSOC format) to create a new SOC (slide.slide).
Also supports parsing the folder structure convention:
  eBSOC/{station}/{name}.xlsx  → BSOC type
  eASOC/{area}/{name}.xlsx    → ASOC type
"""

import base64
import io
import logging
import re

from odoo import models, fields, api, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)


class SocImportWizard(models.TransientModel):
    _name = 'soc.import.wizard'
    _description = 'Import SOC from Excel'

    # ── Upload ──────────────────────────────────────────────────────────────
    excel_file = fields.Binary(string='File Excel (.xlsx)', required=True)
    excel_filename = fields.Char(string='Tên file')

    # ── Classification (user can override after parse) ───────────────────
    soc_name = fields.Char(string='Tên SOC')
    soc_version = fields.Char(string='Version', default='1.0')
    soc_type_id = fields.Many2one('mcd.soc.type', string='Loại SOC')
    soc_area_id = fields.Many2one('mcd.soc.area', string='Khu vực')
    soc_station_id = fields.Many2one(
        'mcd.soc.station', string='Trạm',
        domain="[('area_id', '=', soc_area_id)]"
    )
    document_type = fields.Selection([
        ('permanent', 'Permanent'),
        ('temporary', 'Temporary / LTO'),
    ], string='Loại tài liệu', default='permanent')

    force_update = fields.Boolean(
        string='Ghi đè nếu đã tồn tại',
        help='Nếu SOC cùng tên + station đã tồn tại, ghi đè nội dung cũ.'
    )

    # ── Preview (read-only) ──────────────────────────────────────────────
    preview_html = fields.Html(string='Xem trước', readonly=True)

    # ────────────────────────────────────────────────────────────────────────
    # onchange: parse file on upload for preview/auto-fill
    # ────────────────────────────────────────────────────────────────────────
    @api.onchange('excel_file', 'excel_filename')
    def _onchange_excel_file(self):
        if not self.excel_file:
            return
        try:
            data = self._parse_excel(self.excel_file, self.excel_filename or '')
        except Exception as e:
            self.preview_html = f'<div class="alert alert-danger">Lỗi đọc file: {e}</div>'
            return

        # Auto-fill name from parsed
        if data.get('name') and not self.soc_name:
            self.soc_name = data['name']

        # Auto-detect type from filename prefix
        if not self.soc_type_id:
            fname = (self.excel_filename or '').lower()
            if fname.startswith('easoc') or '/easoc' in fname:
                t = self.env['mcd.soc.type'].search([('code', '=', 'asoc')], limit=1)
                self.soc_type_id = t
            elif fname.startswith('ebsoc') or '/ebsoc' in fname:
                t = self.env['mcd.soc.type'].search([('code', '=', 'bsoc')], limit=1)
                self.soc_type_id = t

        # Auto-detect document type
        service_raw = (data.get('service_type') or '').lower()
        if 'lto' in service_raw or 'temporary' in service_raw:
            self.document_type = 'temporary'
        else:
            self.document_type = 'permanent'

        # Build preview HTML
        items = data.get('items', [])
        rows = ''.join(
            f'<tr>'
            f'<td style="white-space:nowrap">{i+1}</td>'
            f'<td>{it.get("section","")}</td>'
            f'<td>{it.get("name","")}</td>'
            f'<td style="text-align:center">{"⚠️" if it.get("is_critical") else ""}</td>'
            f'</tr>'
            for i, it in enumerate(items)
        )
        self.preview_html = f'''
            <div class="alert alert-info">
                <b>Tên:</b> {data.get("name","?")} &nbsp;|&nbsp;
                <b>Version:</b> {data.get("version","?")} &nbsp;|&nbsp;
                <b>Service:</b> {data.get("service_type","?")} &nbsp;|&nbsp;
                <b>Số mục:</b> {len(items)}
            </div>
            <table class="table table-sm table-bordered table-striped" style="font-size:12px">
                <thead class="table-dark">
                    <tr><th>#</th><th>Section</th><th>Nội dung</th><th>KO</th></tr>
                </thead>
                <tbody>{rows}</tbody>
            </table>
        '''

    # ────────────────────────────────────────────────────────────────────────
    # MAIN ACTION
    # ────────────────────────────────────────────────────────────────────────
    def action_import_soc(self):
        self.ensure_one()
        if not self.excel_file:
            raise UserError(_('Vui lòng chọn file Excel!'))

        try:
            data = self._parse_excel(self.excel_file, self.excel_filename or '')
        except Exception as e:
            raise UserError(_(f'Không thể đọc file Excel:\n{e}'))

        soc_name = self.soc_name or data.get('name') or (self.excel_filename or '').replace('.xlsx', '')
        if not soc_name:
            raise UserError(_('Không xác định được tên SOC. Vui lòng nhập thủ công.'))

        # Check duplicate
        existing = self.env['slide.slide'].search([
            ('is_soc', '=', True),
            ('name', '=', soc_name),
            ('soc_station_id', '=', self.soc_station_id.id if self.soc_station_id else False),
        ], limit=1)

        if existing and not self.force_update:
            raise UserError(_(
                f'SOC "{soc_name}" tại trạm "{self.soc_station_id.name or "N/A"}" đã tồn tại!\n'
                'Bật "Ghi đè nếu đã tồn tại" để cập nhật.'
            ))

        items = data.get('items', [])

        # Build item commands
        item_cmds = []
        for it in items:
            sec = self._get_or_create_section(it.get('section', ''))
            item_cmds.append((0, 0, {
                'section_id': sec.id,
                'sequence': it.get('sequence', 10),
                'name': it['name'],
                'is_critical': it.get('is_critical', False),
            }))

        if existing and self.force_update:
            existing.soc_item_ids.unlink()
            existing.write({
                'soc_version': self.soc_version or data.get('version', '1.0'),
                'soc_type_id': self.soc_type_id.id if self.soc_type_id else existing.soc_type_id.id,
                'soc_area_id': self.soc_area_id.id if self.soc_area_id else existing.soc_area_id.id,
                'soc_station_id': self.soc_station_id.id if self.soc_station_id else existing.soc_station_id.id,
                'document_type': self.document_type,
                'soc_item_ids': item_cmds,
            })
            slide = existing
            msg = f'✅ Đã cập nhật SOC "{soc_name}" với {len(items)} mục.'
        else:
            # Find/create channel
            channel = self.soc_station_id.channel_id if self.soc_station_id else False
            slide = self.env['slide.slide'].create({
                'name': soc_name,
                'is_soc': True,
                'is_locked': True,
                'channel_id': channel.id if channel else False,
                'soc_version': self.soc_version or data.get('version', '1.0'),
                'soc_type_id': self.soc_type_id.id if self.soc_type_id else False,
                'soc_area_id': self.soc_area_id.id if self.soc_area_id else False,
                'soc_station_id': self.soc_station_id.id if self.soc_station_id else False,
                'document_type': self.document_type,
                'slide_category': 'document',
                'soc_item_ids': item_cmds,
            })
            msg = f'✅ Đã tạo SOC "{soc_name}" với {len(items)} mục.'

        # Open the created/updated SOC record using ORM action to get full views list
        slide_action = self.env['ir.actions.act_window']._for_xml_id(
            'M02_P0209_02.action_lnd_soc_template_slide'
        ) if self.env['ir.model.data'].search([
            ('module', '=', 'M02_P0209_02'),
            ('name', '=', 'action_lnd_soc_template_slide')
        ], limit=1) else {}

        if slide_action:
            slide_action.update({
                'res_id': slide.id,
                'view_mode': 'form',
                'views': [(False, 'form')],
                'target': 'current',
            })

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': 'Import SOC thành công!',
                'message': msg,
                'type': 'success',
                'sticky': False,
                'next': slide_action if slide_action else {'type': 'ir.actions.act_window_close'},
            }
        }

    # ────────────────────────────────────────────────────────────────────────
    # EXCEL PARSER  (McDonald's eSOC format)
    # ────────────────────────────────────────────────────────────────────────
    def _parse_excel(self, file_b64, filename):
        """
        Parse McDonald's eSOC Excel file.

        Column layout (0-indexed):
          SOC sheet:
            R1  (row 1)  : Col 0 = station/SOC name, Col 10 = "Station Observation Checklist..."
            R7  (row 7)  : METADATA
                            Col 0 = SOC Code
                            Col 1 = SOC Type  (BSOC / ASOC / Quiz)
                            Col 3 = Document type  (Permanent / LTO / ...)
                            Col 5 = Area  (Kitchen / Service)
                            Col 7 = Station sub-area  (Production / McCafe / etc.)
            R9  (row 9)  : Col 0 = Version  (e.g. 082023)
            R18+ (row 18+): Content rows
                            Col 1 (B) non-empty + Col 2 (C) empty → Section header
                            Col 2 (C) non-empty                   → Checklist item

        Returns dict: {
          'name': str,
          'soc_code': str,
          'soc_type': str,        # 'BSOC' | 'ASOC' | 'Quiz'
          'doc_type': str,        # 'permanent' | 'temporary'
          'area_name': str,       # 'Kitchen' | 'Service'
          'station_name': str,    # 'Production' | 'McCafe' | etc.
          'version': str,
          'items': [
            {'sequence': int, 'section': str, 'name': str, 'is_critical': bool}
          ]
        }
        """
        try:
            import openpyxl
        except ImportError:
            raise UserError(_('Thiếu thư viện openpyxl. Chạy: pip install openpyxl'))

        raw = base64.b64decode(file_b64)

        # ── PRE-PROCESS: strip xlsx error cells before openpyxl parses them ──
        # openpyxl raises ValueError("Invalid cell value... #DIV/0!") INSIDE
        # load_workbook() during XML parsing — no try/except can catch this.
        # Fix: edit the xlsx ZIP in memory, removing all <c t="e">...</c> nodes.
        import zipfile, re as _re

        clean_buf = io.BytesIO()
        with zipfile.ZipFile(io.BytesIO(raw), 'r') as zin, \
             zipfile.ZipFile(clean_buf, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if (item.filename.startswith('xl/worksheets/') and
                        item.filename.endswith('.xml')):
                    xml = data.decode('utf-8', errors='replace')
                    # Remove error cells: <c ... t="e" ...>...</c>
                    xml = _re.sub(
                        r'<c\b[^>]*\bt="e"\b[^>]*>.*?</c>',
                        '', xml, flags=_re.DOTALL
                    )
                    xml = _re.sub(
                        r'<c\b[^>]*\bt="e"\b[^>]*/>', '', xml
                    )
                    data = xml.encode('utf-8')
                zout.writestr(item, data)

        wb = openpyxl.load_workbook(clean_buf, data_only=False)


        ws = wb['SOC'] if 'SOC' in wb.sheetnames else wb.active

        # ── Safe row reader (handles any remaining exceptions) ──
        def _cell(row_tuple, col_idx, default=''):
            try:
                v = row_tuple[col_idx] if col_idx < len(row_tuple) else None
                if v is None:
                    return default
                s = str(v).strip()
                # Skip formula strings – treat as empty
                if s.startswith('='):
                    return default
                return s
            except Exception:
                return default

        # Read all rows into a list (max 20 cols)
        MAX_COLS = 20
        all_rows = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=MAX_COLS):
            safe_row = []
            for c in row:
                try:
                    v = c.value
                    # Skip formula values (start with '=') — not needed for our parser
                    if isinstance(v, str) and v.startswith('='):
                        safe_row.append(None)
                    else:
                        safe_row.append(v)
                except Exception:
                    safe_row.append(None)
            all_rows.append(safe_row)


        result = {
            'name': '',
            'soc_code': '',
            'soc_type': '',
            'doc_type': 'permanent',
            'area_name': '',
            'station_name': '',
            'version': '',
            'items': [],
        }

        # ── Row 1 → SOC name (Col A) ────────────────────────────────────
        if all_rows:
            result['name'] = _cell(all_rows[0], 0)

        # ── Row 7 (index 6) → metadata ──────────────────────────────────
        if len(all_rows) >= 7:
            meta = all_rows[6]
            result['soc_code']    = _cell(meta, 0)
            result['soc_type']    = _cell(meta, 1)    # BSOC / ASOC / Quiz
            doc_raw               = _cell(meta, 3).lower()
            result['doc_type']    = 'temporary' if 'lto' in doc_raw or 'temp' in doc_raw else 'permanent'
            result['area_name']   = _cell(meta, 5)    # Kitchen / Service
            result['station_name'] = _cell(meta, 7)   # Production / McCafe / etc.

        # ── Row 9 (index 8) → version ───────────────────────────────────
        if len(all_rows) >= 9:
            ver_raw = _cell(all_rows[8], 0)
            if ver_raw:
                result['version'] = ver_raw

        # ── Fallback: name from filename if R1 was empty ─────────────────
        if not result['name'] and filename:
            result['name'] = re.sub(r'[-_]?v?\d{4,}\.xlsx$', '', filename, flags=re.I).strip()
            result['name'] = re.sub(r'^eASOC\s+|^eBSOC\s+', '', result['name'], flags=re.I).strip()

        # ── Content rows: start from row 17 (index 16) ──────────────────
        # Skip instruction rows (Bước 1, 2, 3... rows before row 17)
        CONTENT_START = 16  # row index (0-based) ≈ row 17 in Excel
        current_section = ''
        seq = 10

        for row_data in all_rows[CONTENT_START:]:
            col_b = _cell(row_data, 1)   # Col B – section label
            col_c = _cell(row_data, 2)   # Col C – item description
            col_d = _cell(row_data, 3)   # Col D – sub-data / detail

            # Skip empty rows
            if not col_b and not col_c:
                continue

            # Detect section header: col B has text, col C is empty
            # Also treat table/data rows (col D has food/time data) as items under section
            if col_b and not col_c:
                current_section = col_b.rstrip(':').strip()
                continue

            # Checklist item: col C has content
            if col_c:
                desc = col_c.strip()

                # Skip very short or purely numeric entries
                if not desc or len(desc) < 3:
                    continue
                # Skip header-like rows that slipped through
                if desc in ('Mã', 'Loại tài liệu', 'Phân loại', 'Khu vực', 'Thực phẩm',
                            'Thời gian', 'Nhiệt độ'):
                    continue

                # Check for K.O marker (look in cols 14-19 for 'x' or 'k.o' or 'ko')
                is_critical = any(
                    str(row_data[ci]).strip().upper() in ('X', 'K.O', 'KO', 'KO!')
                    for ci in range(14, min(20, len(row_data)))
                    if row_data[ci] is not None
                )

                result['items'].append({
                    'sequence': seq,
                    'section': current_section,
                    'name': desc,
                    'is_critical': is_critical,
                })
                seq += 10

        return result

    # ────────────────────────────────────────────────────────────────────────
    # AUTO-FILL from parsed metadata (onchange override)
    # ────────────────────────────────────────────────────────────────────────
    @api.onchange('excel_file', 'excel_filename')
    def _onchange_excel_file(self):
        if not self.excel_file:
            return
        try:
            data = self._parse_excel(self.excel_file, self.excel_filename or '')
        except Exception as e:
            self.preview_html = f'<div class="alert alert-danger"><b>Lỗi đọc file:</b> {e}</div>'
            return

        # Auto-fill name
        if data.get('name') and not self.soc_name:
            self.soc_name = data['name']

        # Auto-fill version
        if data.get('version'):
            self.soc_version = data['version']

        # Auto-detect SOC type
        if not self.soc_type_id and data.get('soc_type'):
            code_map = {'BSOC': 'bsoc', 'ASOC': 'asoc', 'Quiz': 'quiz', 'QUIZ': 'quiz'}
            code = code_map.get(data['soc_type'].upper(), data['soc_type'].lower())
            t = self.env['mcd.soc.type'].search([('code', '=', code)], limit=1)
            if not t:
                t = self.env['mcd.soc.type'].search([('name', 'ilike', data['soc_type'])], limit=1)
            self.soc_type_id = t

        # Auto-detect area
        if not self.soc_area_id and data.get('area_name'):
            area = self.env['mcd.soc.area'].search([('name', 'ilike', data['area_name'])], limit=1)
            self.soc_area_id = area

        # Auto-detect station
        if not self.soc_station_id and data.get('station_name'):
            dom = [('name', 'ilike', data['station_name'])]
            if self.soc_area_id:
                dom.append(('area_id', '=', self.soc_area_id.id))
            stn = self.env['mcd.soc.station'].search(dom, limit=1)
            self.soc_station_id = stn

        # Doc type
        self.document_type = data.get('doc_type', 'permanent')

        # Build preview
        items = data.get('items', [])
        rows = ''.join(
            f'<tr>'
            f'<td>{i+1}</td>'
            f'<td style="color:#888">{it.get("section","")}</td>'
            f'<td>{it.get("name","")}</td>'
            f'<td style="text-align:center">{"⚠️" if it.get("is_critical") else ""}</td>'
            f'</tr>'
            for i, it in enumerate(items)
        )
        self.preview_html = f'''
            <div class="alert alert-info mb-2">
                <b>Tên:</b> {data.get("name","?")} &nbsp;|&nbsp;
                <b>Mã:</b> {data.get("soc_code","?")} &nbsp;|&nbsp;
                <b>Loại:</b> {data.get("soc_type","?")} &nbsp;|&nbsp;
                <b>Phiên bản:</b> {data.get("version","?")} &nbsp;|&nbsp;
                <b>Khu vực:</b> {data.get("area_name","?")} &nbsp;|&nbsp;
                <b>Trạm:</b> {data.get("station_name","?")} &nbsp;|&nbsp;
                <b>Loại tài liệu:</b> {data.get("doc_type","?")} &nbsp;|&nbsp;
                <b>Số mục:</b> {len(items)}
            </div>
            <table class="table table-sm table-bordered table-striped" style="font-size:12px">
                <thead class="table-dark">
                    <tr><th>#</th><th>Section</th><th>Nội dung</th><th>KO</th></tr>
                </thead>
                <tbody>{rows}</tbody>
            </table>
        '''

    # ────────────────────────────────────────────────────────────────────────
    # HELPERS
    # ────────────────────────────────────────────────────────────────────────
    def _get_or_create_section(self, name):
        name = (name or '').strip() or 'Chung'
        sec = self.env['mcd.soc.section'].search([('name', '=', name)], limit=1)
        if not sec:
            sec = self.env['mcd.soc.section'].create({'name': name})
        return sec

